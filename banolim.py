# ----------------- 라이브러리 임포트 -----------------
import streamlit as st
import pandas as pd
import requests
import xml.etree.ElementTree as ET
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from io import BytesIO
import re

# ----------------- 고정 유해성 항목 순서 -----------------
HAZARD_ORDER = [
    '#', '물질명칭', 'CAS No.', '결과없음',
    '발암성', '생식독성', '생식세포 변이원성', 'CMR',
    '급성 독성(경구)', '급성 독성(경피)', '급성 독성(흡입)',
    '흡인 유해성', '피부 부식성/피부 자극성', '심한 눈 손상성/눈 자극성',
    '호흡기 과민성', '피부 과민성', '특정표적장기 독성(1회 노출)',
    '특정표적장기 독성(반복 노출)', '급성 수생환경 유해성', '만성 수생환경 유해성',
    '폭발성 물질', '자기반응성 물질', '유기과산화물', '산화성 가스',
    '산화성 액체', '산화성 고체', '인화성 가스', '인화성 에어로졸',
    '인화성 액체', '인화성 고체', '자연발화성 액체', '자연발화성 고체',
    '물반응성 물질', '고압가스', '자기발열성 물질', '금속부식성 물질',
    'TWA', 'STEL', '증기압', '개정일',
    '관리대상유해물질', '특별관리물질', 
    '작업환경측정대상물질', #0930
    '특수건강진단대상물질',
    '노출기준설정물질','허용기준설정물질','금지물질','제한물질','유독물질', #0930
    '허가물질','사고대비물질','중점관리물질','위험물','독성가스', #0930
    '인체급성유해성물질', '인체만성유해성물질'
]


# ----------------- CMR 최고 등급 판별 함수 -----------------
def get_highest_cmr_grade(grades):
    priority = {'1A': 0, '1B': 1, '2': 2}
    best = None
    for g in grades:
        if g in priority:
            if best is None or priority[g] < priority[best]:
                best = g
    return best

# ----------------- TWA, STEL 조회 함수 -----------------
def query_twa_stel(service_key, chem_id):
    try:
        res = requests.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail08',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)

        twa = ''
        stel = ''

        for item in root.findall('.//item'):
            name_twa = item.findtext('msdsItemNameKor')
            if name_twa == '국내규정':
                detail = item.findtext('itemDetail')
                if detail:
                    parts = [p.strip() for p in detail.split('|') if p.strip()]
                    for part in parts:
                        if part.startswith('TWA'):
                            match = re.search(r'TWA\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
                            if match:
                                twa = match.group(1).strip()
                        elif part.startswith('STEL'):
                            match = re.search(r'STEL\s*[:]?\s*([\d\.]+\s*ppm(?:\([^)]*\))?)', part)
                            if match:
                                stel = match.group(1).strip()

        return twa, stel
    except:
        return '', ''

# ----------------- 증기압 조회 함수 -----------------
def query_vapor_pressure(service_key, chem_id):
    import html
    
    try:
        res = requests.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail09',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)

        for item in root.findall('.//item'):
            name = item.findtext('msdsItemNameKor')
            if name and name.strip() == '증기압':
                detail = item.findtext('itemDetail')
                detail = html.unescape(detail)
                if detail and detail.strip():
                    cleaned = ''.join(detail.split())  # 공백 제거
                    # '|' 또는 '※' 이후의 모든 내용을 제거
                    cleaned = re.split(r'\|+|※+', cleaned)[0]
                    return cleaned
    except:
        pass
    return ''


# ----------------- 개정일 조회 함수 -----------------
def query_revision_date(service_key, chem_id):
    try:
        res = requests.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail16',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)
        for item in root.findall('.//item'):
            if item.findtext('msdsItemNameKor') == '최종 개정일자':
                detail = item.findtext('itemDetail')
                return detail.strip() if detail else ''
    except:
        return ''
    return ''

# --- '관리대상유해물질', '특별관리물질', '특수건강진단대상물질', '인체급성유해성물질', '인체만성유해성물질' 조회 함수 ---
# 0930 수정
def has_keyword(detail, keyword):
    """
    itemDetail 문자열에서 '|'로 구분된 항목 중 keyword를 포함한 항목이 있는지 검사
    """
    if not detail:
        return False
    items = [d.strip() for d in detail.split('|')]
    return any(keyword in item for item in items)

def query_detail15(service_key, chem_id):
    try:
        res = requests.get(
            'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail15',
            params={'serviceKey': service_key, 'chemId': chem_id},
            timeout=10
        )
        res.encoding = 'utf-8'
        root = ET.fromstring(res.text)

        result = {
            '관리대상유해물질': False,
            '특별관리물질': False,
            '작업환경측정대상물질': False, #0930
            '특수건강진단대상물질': False,
            #0930-----------
            '노출기준설정물질': False,
            '허용기준설정물질': False,
            '금지물질': False,
            '제한물질': False,
            '유독물질': False,
            '허가물질': False,
            '사고대비물질': False,
            '중점관리물질': False,
            '위험물': False,
            '독성가스': False,
            #--------------0930
            '인체급성유해성물질': False,
            '인체만성유해성물질': False
        }

        for item in root.findall('.//item'):
            code = item.findtext('msdsItemCode')
            detail = item.findtext('itemDetail') or ''

            #0930----------------------------------------
            if code in ['002', 'O04', 'O12', 'O06']:  # 산업안전보건법 / 화학물질관리법 / 등록평가법 / 위험물안전관리법
                if has_keyword(detail, '금지물질'):
                    result['금지물질'] = True
                if has_keyword(detail, '제한물질'):
                    result['제한물질'] = True
                if has_keyword(detail, '유독물질'):
                    result['유독물질'] = True
                if has_keyword(detail, '허가물질'):
                    result['허가물질'] = True
                if has_keyword(detail, '사고대비물질'):
                    result['사고대비물질'] = True
                if has_keyword(detail, '중점관리물질'):
                    result['중점관리물질'] = True
                if has_keyword(detail, '위험물'):
                    result['위험물'] = True
                if has_keyword(detail, '독성가스'):
                    result['독성가스'] = True
            #0930----------------------------------------

            if code == 'O02':  # 산업안전보건법에 의한 규제
                if has_keyword(detail, '관리대상유해물질'):
                    result['관리대상유해물질'] = True
                if has_keyword(detail, '특별관리물질'):
                    result['특별관리물질'] = True
                if has_keyword(detail, '특수건강진단대상물질'):
                    result['특수건강진단대상물질'] = True
                #0930-----------------------------------------
                if has_keyword(detail, '작업환경측정대상물질'):
                    result['작업환경측정대상물질'] = True
                if has_keyword(detail, '노출기준설정물질'):
                    result['노출기준설정물질'] = True
                if has_keyword(detail, '허용기준설정물질'):
                    result['허용기준설정물질'] = True
                #-----------------------------------------0930

            elif code in ['O04', 'O12']:  # 화학물질관리법 / 등록평가법
                if has_keyword(detail, '인체급성유해성물질'):
                    result['인체급성유해성물질'] = True
                if has_keyword(detail, '인체만성유해성물질'):
                    result['인체만성유해성물질'] = True
                    
        print(f'detail15 result({chem_id})\n{result}')

        return result

    except Exception as e:
        # 예외 발생 시 모두 False로 반환
        return {
            '관리대상유해물질': False,
            '특별관리물질': False,
            '작업환경측정대상물질': False, #0930
            '특수건강진단대상물질': False,
            #0930-----------
            '노출기준설정물질': False,
            '허용기준설정물질': False,
            '금지물질': False,
            '제한물질': False,
            '유독물질': False,
            '허가물질': False,
            '사고대비물질': False,
            '중점관리물질': False,
            '위험물': False,
            '독성가스': False,
            #-------------0930
            '인체급성유해성물질': False,
            '인체만성유해성물질': False            
        }
        
# ----------------- CAS 정보 조회 함수 -----------------
def query_cas_info(data_rows, service_key):
    results = []
    res_detail15 = []
    unknown_columns = set()
    progress = st.progress(0)

    for idx, row in enumerate(data_rows.itertuples(index=False), start=1):
        cas = row[2]
        id_num = row[0]
        name = row[1]
        result = {'#': id_num, '물질명칭': name, 'CAS No.': cas}

        try:
            res_id = requests.get(
                'https://msds.kosha.or.kr/openapi/service/msdschem/chemlist',
                params={'serviceKey': service_key, 'searchWrd': cas, 'searchCnd': 1},
                timeout=10
            )
            res_id.encoding = 'utf-8'
            chem_id = ET.fromstring(res_id.text).findtext('.//chemId')

            if not chem_id:
                result['결과없음'] = '공단 MSDS 없음'
            else:
                res_detail = requests.get(
                    'https://msds.kosha.or.kr/openapi/service/msdschem/chemdetail02',
                    params={'serviceKey': service_key, 'chemId': chem_id},
                    timeout=10
                )
                res_detail.encoding = 'utf-8'
                root = ET.fromstring(res_detail.text)

                b02_detail = next(
                    (item.findtext('itemDetail')
                     for item in root.findall('.//item')
                     if item.findtext('msdsItemCode') == 'B02'),
                    None
                )

                if b02_detail is None or b02_detail.strip() == '' or b02_detail.strip() == '자료없음':
                    result['결과없음'] = '자료 없음'
                else:
                    merged = defaultdict(list)
                    inhalation_labels = [
                        '급성 독성(흡입)', '급성 독성(흡입: 가스)',
                        '급성 독성(흡입: 분진/미스트)', '급성 독성(흡입: 증기)'
                    ]
                    inhalation_entries = []
                    cmr_map = {'발암성': [], '생식독성': [], '생식세포 변이원성': []}

                    for entry in b02_detail.split('|'):
                        if ':' in entry and '자료없음' not in entry:
                            k, v = map(str.strip, entry.rsplit(':', 1))
                            v = v.replace('구분', '').strip()
                            if k in inhalation_labels:
                                label = k.replace('급성 독성(', '').replace(')', '')
                                inhalation_entries.append(f"{v}({label})")
                            elif k in cmr_map:
                                # 0930 : CAS가 에탄올이고 발암성일 경우 cmr_map에도 넣지 않음 
                                if not (cas == '64-17-5' and k == '발암성'):
                                    cmr_map[k].append(v)
                                    if v not in merged[k]:
                                        merged[k].append(v)
                            else:
                                if v not in merged[k]:
                                    merged[k].append(v)

                    if inhalation_entries:
                        result['급성 독성(흡입)'] = '\n'.join(inhalation_entries)

                    for k, v_list in merged.items():
                        # 예외 처리: CAS No가 64-17-5(에탄올)이고, 항목이 발암성이면 무시
                        if cas == '64-17-5' and k == '발암성':
                            continue  # 이 항목은 기록하지 않음

                        result[k] = '\n'.join(v_list)
                        if k not in HAZARD_ORDER:
                            unknown_columns.add(k)

                    cmr_grades = []
                    for values in cmr_map.values():
                        for v in values:
                            if v in ['1A', '1B', '2']:
                                cmr_grades.append(v)
                    best_grade = get_highest_cmr_grade(cmr_grades)
                    if best_grade:
                        result['CMR'] = best_grade

                result['TWA'], result['STEL'] = query_twa_stel(service_key, chem_id)
                result['증기압'] = query_vapor_pressure(service_key, chem_id)
                result['개정일'] = query_revision_date(service_key, chem_id)
                
                res_detail15 = query_detail15(service_key, chem_id)
                if res_detail15['관리대상유해물질'] :
                    result['관리대상유해물질'] = '▣'
                if res_detail15['특별관리물질'] :
                    result['특별관리물질'] = '▣'
                if res_detail15['특수건강진단대상물질'] :
                    result['특수건강진단대상물질'] = '▣'
                if res_detail15['인체급성유해성물질'] :
                    result['인체급성유해성물질'] = '▣'
                if res_detail15['인체만성유해성물질'] :
                    result['인체만성유해성물질'] = '▣'
                #0930--------------------------------
                if res_detail15['작업환경측정대상물질'] :
                    result['작업환경측정대상물질'] = '▣'
                if res_detail15['노출기준설정물질'] :
                    result['노출기준설정물질'] = '▣'
                if res_detail15['허용기준설정물질'] :
                    result['허용기준설정물질'] = '▣'
                if res_detail15['금지물질'] :
                    result['금지물질'] = '▣'
                if res_detail15['제한물질'] :
                    result['제한물질'] = '▣'
                if res_detail15['유독물질'] :
                    result['유독물질'] = '▣'
                if res_detail15['허가물질'] :
                    result['허가물질'] = '▣'
                if res_detail15['사고대비물질'] :
                    result['사고대비물질'] = '▣'
                if res_detail15['중점관리물질'] :
                    result['중점관리물질'] = '▣'
                if res_detail15['위험물'] :
                    result['위험물'] = '▣'
                if res_detail15['독성가스'] :
                    result['독성가스'] = '▣'
                #--------------------------------0930

        except Exception as e:
            result['결과없음'] = f'조회 오류: {str(e)}'


        results.append(result)
        progress.progress(min(idx / len(data_rows), 1.0))

    df = pd.DataFrame(results)
    return df, sorted(list(unknown_columns))


# ----------------- Streamlit 앱 실행 -----------------
import os

st.set_page_config(page_title="화학물질 유해성 정보 수집기", layout="wide")
st.title("📋 화학물질 유해성 정보 수집기 v.250930_1")

SERVICE_KEY = 'MJFEGDzjkGr4Rg4pQtOxcYT%2BxteNCe0HuK0PUWKt%2B4hZHqYk%2BpNIf3RwocbhI1twsbNknwMur9m0fcPZir9jyg%3D%3D'

# ----------------- 세션 초기화 -----------------
if 'processed' not in st.session_state:
    st.session_state.processed = False
if 'result_file' not in st.session_state:
    st.session_state.result_file = None
if 'uploader_key' not in st.session_state:
    st.session_state.uploader_key = 0  # file_uploader 초기 키

# ----------------- 파일 업로드 -----------------
uploaded_file = st.file_uploader(
    "📎 엑셀 파일을 업로드 하세요! (입력파일명: A.xlsx, A: 회사명)", 
    type="xlsx", 
    key=f"file_uploader_{st.session_state.uploader_key}"
)

# ----------------- 처리 로직 -----------------
if uploaded_file and not st.session_state.processed:
    wb = load_workbook(uploaded_file)
    ws = wb.active

    raw_df = pd.read_excel(uploaded_file, header=None)
    header_row_full = raw_df.iloc[0].tolist()
    data_rows = raw_df[1:].copy()
    data_rows.columns = header_row_full  # 전체 열 이름을 유지

    #header_row = header_row_full[:45]  # ✅ A열~AS열(1~45열)만 제목행 검사
    header_row = header_row_full[:56]  # 0930 : A열~BD열(1~56열)만 제목행 검사
    current_headers = set(header_row)
    #print('[current_headers]\n', current_headers)
    expected_headers = set(HAZARD_ORDER)

    unexpected_headers = [h for h in header_row if h not in expected_headers]
    #print('[unexpected_headers]\n', unexpected_headers)
    missing_headers = [h for h in HAZARD_ORDER if h not in current_headers]
    #print('[missing_headers]\n', missing_headers)
    
    
    if unexpected_headers or missing_headers:
        st.error("❗제목행(A~AN열)에 오류가 있습니다. 유해성 정보를 조회하지 않습니다.")

        if unexpected_headers:
            st.markdown("### 🚫 예기치 않은 열 제목")
            for col in unexpected_headers:
                st.markdown(f"- `{col}`")

        if missing_headers:
            st.markdown("### ⚠️ 누락된 필수 항목")
            for col in missing_headers:
                st.markdown(f"- `{col}`")

        st.stop()

    # ✅ 유해성 정보 조회
    required_cols = {'#', '물질명칭', 'CAS No.'}
    if not required_cols.issubset(set(data_rows.columns)):
        st.error("필수 열('#', '물질명칭', 'CAS No.')이 누락되어 있습니다.")
        st.stop()
        
    # CAS No. 전처리 (앞자리 0 제거)
    data_rows['CAS No.'] = data_rows['CAS No.'].astype(str).str.lstrip('0')

    data_rows = data_rows[['#', '물질명칭', 'CAS No.']].copy()
    hazard_df, unknown_hazards = query_cas_info(data_rows, SERVICE_KEY)

    # ✅ 컬럼명 → 엑셀 열 인덱스 매핑
    col_name_to_idx = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}

    # ✅ 결과 입력 + 셀 정렬 설정
    for r_idx, row in hazard_df.iterrows():
        excel_row = r_idx + 2  # 엑셀은 1-based, 데이터는 2행부터 시작
        for col_name in HAZARD_ORDER:
            if col_name in col_name_to_idx:
                col_idx = col_name_to_idx[col_name]
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = row.get(col_name, '')
                
                # 셀 정렬: CAS No. 등은 가운데 정렬, 그 외는 위쪽 정렬
                if col_name in ['#', 'CAS No.', '결과없음', '개정일',
                                '관리대상유해물질', '특별관리물질', '특수건강진단대상물질', '인체급성유해성물질', '인체만성유해성물질',
                                #0930---------------------------
                                '작업환경측정대상물질', '노출기준설정물질','허용기준설정물질','금지물질','제한물질','유독물질',
                                '허가물질','사고대비물질','중점관리물질','위험물','독성가스']:
                                #---------------------------0930
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    

    #=========================================================================#
    # 표2 생성
    #=========================================================================#
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Border, Side, Font, Alignment
    import math
    import re

    # -------------------- 설정 --------------------
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    default_font = Font(name='Noto Sans KR', size=10)

    # -------------------- 등급 판별 기준 --------------------
    grade_to_label = {
        '1': '구분1',
        '1A': '1A',
        '1B': '1B',
        '2': '구분2',
        '3': '구분3',
        '4': '구분4'
    }
    grade_priority = {'1': 0, '1A': 1, '1B': 2, '2': 3, '3': 4, '4': 5}

    # -------------------- 등급 추출 함수 --------------------
    def extract_most_severe_grade(cell_value):
        if not isinstance(cell_value, str):
            return None
        parts = re.split(r'[\n|,]+', cell_value)
        found = []
        for part in parts:
            match = re.match(r'^\s*(1A|1B|1|2|3|4)\b', part.strip())
            if match:
                found.append(match.group(1))
        if not found:
            return None
        return min(found, key=lambda g: grade_priority[g])

    # -------------------- 기본 정보 --------------------
    #hazard_cols = HAZARD_ORDER[4:-9]
    hazard_cols = HAZARD_ORDER[4:-20]  # 0930 발암성~금속부식성물질
    hazard_start_col = 4  # D열 = openpyxl 기준 1-based
    start_row = 2
    end_row = start_row + len(hazard_df) - 1
    summary_start_row = end_row + 2

    # -------------------- 제목행 --------------------
    ws[f"D{summary_start_row}"] = "유해성"
    ws[f"D{summary_start_row}"].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws[f"D{summary_start_row}"].font = default_font
    ws[f"D{summary_start_row}"].border = thin_border

    for idx, col_name in enumerate(hazard_cols):
        col_letter = get_column_letter(hazard_start_col + idx + 1)
        cell = ws[f"{col_letter}{summary_start_row}"]
        cell.value = col_name
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = default_font
        cell.border = thin_border

    # -------------------- 라벨 --------------------
    row_labels = [
        '구분1', '1A', '1B', '구분2', '구분3', '구분4', '기타구분',
        '유해물질수', '분석물질수', '유해물질비율'
    ]

    # -------------------- 분석물질수 --------------------
    analyzed_count = sum(
        1 for r in range(start_row, end_row + 1)
        if ws.cell(row=r, column=HAZARD_ORDER.index('결과없음') + 1).value != '공단 MSDS 없음'
    )

    # -------------------- 표2 생성 --------------------
    summary_data = []

    for hazard in hazard_cols:
        col_idx = HAZARD_ORDER.index(hazard) + 1
        count_map = {label: 0 for label in row_labels[:-3]}  # 유해물질수, 분석물질수, 비율 제외

        for r in range(start_row, end_row + 1):
            val = ws.cell(row=r, column=col_idx).value

            if val is None or (isinstance(val, float) and math.isnan(val)) or str(val).strip() == '':
                continue

            most_severe = extract_most_severe_grade(str(val))

            if most_severe in grade_to_label:
                label = grade_to_label[most_severe]
                count_map[label] += 1
            #elif most_severe is None:
            #    continue
            else:
                count_map['기타구분'] += 1

        count_map['유해물질수'] = sum(count_map[label] for label in row_labels[:7])
        count_map['분석물질수'] = analyzed_count
        count_map['유해물질비율'] = f"{round((count_map['유해물질수'] / analyzed_count) * 100)}%" if analyzed_count else "0%"

        for i, label in enumerate(row_labels):
            if len(summary_data) <= i:
                summary_data.append([])
            summary_data[i].append(count_map[label])

    # -------------------- 표2 입력 --------------------
    for row_offset, (label, row_values) in enumerate(zip(row_labels, summary_data), start=1):
        row_num = summary_start_row + row_offset

        label_cell = ws.cell(row=row_num, column=hazard_start_col)
        label_cell.value = label
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        label_cell.font = default_font
        label_cell.border = thin_border

        for col_offset, value in enumerate(row_values):
            col_num = hazard_start_col + col_offset + 1
            cell = ws.cell(row=row_num, column=col_num)
            
            
            if label in ('유해물질수', '유해물질비율') :
                # 유해물질수와 비율은 그대로 기록
                cell.value = value
            else:
                # 나머지는 0이면 빈 셀
                cell.value = value if value != 0 else None
                                    
            #cell.value = value
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = default_font
            cell.border = thin_border
    #=========================================================================#

    

    #=========================================================================#
    # 표3 생성
    #=========================================================================#
    # ✅ 표1 AO~BA열은 41~58열 (엑셀 기준: AO~BA)
    summary_titles = [
        '관리대상유해물질', '특별관리물질', 
        '작업환경측정대상물질', #0930    
        '특수건강진단대상물질',
        #0930------------------------------
        '노출기준설정물질','허용기준설정물질','금지물질','제한물질','유독물질',
        '허가물질','사고대비물질','중점관리물질','위험물','독성가스',
        #------------------------------0930
        '인체급성유해성물질', '인체만성유해성물질', 
        '유독물질2', '제한물질2', '금지물질2', '허가물질2', '사고대비물질2',
        '중점관리물질2', '금지·허가물질2', '노출·허용기준물질2', '직업환경측정물질등2',
        '위험물2', '독성가스2'
    ]
    
    # 표3의 컬럼 인덱스 (엑셀 기준 41~58)
    summary_start_col = 41
    summary_end_col = summary_start_col + len(summary_titles) - 1

    # 전체 물질 수 (결과없음이 '공단 MSDS 없음'이 아닌 것들)
    start_row = 2
    end_row = start_row + len(hazard_df) - 1
    analyzed_count = 0
    col_result_idx = HAZARD_ORDER.index('결과없음') + 1

    for r in range(start_row, end_row + 1):
        val = ws.cell(row=r, column=col_result_idx).value
        if val != '공단 MSDS 없음':
            analyzed_count += 1

    # 표3 시작 행 (표2 끝 기준 + 2)
    table3_start_row = end_row + 2
    table3_row_labels = ['규제물질', '물질 수', '물질 비율']

    # ✅ 표3 제목 셀 (AN열)
    ws.cell(row=table3_start_row, column=summary_start_col - 1, value="규제물질").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=table3_start_row, column=summary_start_col - 1, value="규제물질").font = default_font
    ws.cell(row=table3_start_row, column=summary_start_col - 1).border = thin_border
    ws.cell(row=table3_start_row + 1, column=summary_start_col - 1, value="물질 수").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=table3_start_row + 1, column=summary_start_col - 1, value="물질 수").font = default_font
    ws.cell(row=table3_start_row + 1, column=summary_start_col - 1).border = thin_border
    ws.cell(row=table3_start_row + 2, column=summary_start_col - 1, value="물질 비율").alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=table3_start_row + 2, column=summary_start_col - 1, value="물질 비율").font = default_font
    ws.cell(row=table3_start_row + 2, column=summary_start_col - 1).border = thin_border

    # ✅ 표3 열 제목 (summary_titles)
    for idx, col_name in enumerate(summary_titles):
        col_letter = get_column_letter(summary_start_col + idx)
        cell = ws.cell(row=table3_start_row, column=summary_start_col + idx, value=col_name)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = default_font
        cell.border = thin_border

    # ✅ 각 열별 ▣ 개수 세기 (물질 수), 비율 계산
    for idx in range(len(summary_titles)):
        col_idx = summary_start_col + idx
        count = 0
        for r in range(start_row, end_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if str(val).strip() == '▣':
                count += 1

        # 물질 수 입력
        cell = ws.cell(row=table3_start_row + 1, column=col_idx, value=count)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = default_font
        cell.border = thin_border

        # 물질 비율 입력
        ratio = f"{round((count / analyzed_count) * 100)}%" if analyzed_count else "0%"
        cell = ws.cell(row=table3_start_row + 2, column=col_idx, value=ratio)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = default_font
        cell.border = thin_border          
    #=========================================================================#


    #=========================================================================#
    # 표4 생성
    #=========================================================================#
    from openpyxl.utils import get_column_letter
    from collections import Counter

    # ----------------- 엑셀 시트에서 열 인덱스 찾기 -----------------
    header_row = 1
    col_idx_in = None
    col_idx_use = None

    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        if header == '연간입고량':
            col_idx_in = col
        elif header == '연간사용·판매량':
            col_idx_use = col

    if col_idx_in is None or col_idx_use is None:
        raise ValueError("'연간입고량' 또는 '연간사용·판매량' 열이 존재하지 않습니다.")

    # ----------------- 표3 마지막 위치 기준으로 시작행 설정 -----------------
    table3_end_row = table3_start_row + 2  # 표3은 총 3행
    table4_start_row = table3_end_row + 2  # 표3 끝 + 2줄 띄움
    #table4_start_col = 56  # BD열 = 56
    table4_start_col = 67  # 0930 : BO열 = 67

    # ----------------- 표4 열 제목 -----------------
    headers = ['중량(톤/년) 또는 부피단위(㎥/년)', '연간입고량', '연간사용·판매량']
    for idx, header in enumerate(headers):
        col_letter = get_column_letter(table4_start_col + idx)
        cell = ws[f"{col_letter}{table4_start_row}"]
        cell.value = header
        cell.font = default_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # ----------------- 사용량 구분 값 및 범위 -----------------
    usage_levels = [str(i) for i in range(1, 11)]
    usage_map = {
        '1': "0.1미만",
        '2': "0.1~0.5",
        '3': "0.5~1.0",
        '4': "1~2.5",
        '5': "2.5~5.0",
        '6': "5~20",
        '7': "20~200",
        '8': "200~1,000",
        '9': "1,000~5,000",
        '10': "5,000이상",
        1: "0.1미만",
        2: "0.1~0.5",
        3: "0.5~1.0",
        4: "1~2.5",
        5: "2.5~5.0",
        6: "5~20",
        7: "20~200",
        8: "200~1,000",
        9: "1,000~5,000",
        10: "5,000이상",
    }
    usage_descriptions = [
        "0.1미만",         # 1
        "0.1~0.5",         # 2
        "0.5~1.0",         # 3
        "1~2.5",           # 4
        "2.5~5.0",         # 5
        "5~20",            # 6
        "20~200",          # 7
        "200~1,000",       # 8
        "1,000~5,000",     # 9
        "5,000이상"        # 10
    ]
    start_row = 2
    end_row = start_row + len(hazard_df) - 1
    
    # 표1의 두 열 값을 설명 문자열로 변환하여 시트에 반영
    def normalize_to_desc(v):
        if v is None or (isinstance(v, float) and math.isnan(v)) or str(v).strip() == '':
            return None
        key = v if isinstance(v, int) else str(v).strip()
        return usage_map.get(key, None)

    # ----------------- 개수 카운트 -----------------
    incoming_counter = Counter()
    usage_counter = Counter()

    for r in range(start_row, end_row + 1):
        # 셀 값 불러오기
        raw_in = ws.cell(row=r, column=col_idx_in).value
        raw_use = ws.cell(row=r, column=col_idx_use).value

        # 설명 문자열로 변환
        desc_in = normalize_to_desc(raw_in)
        desc_use = normalize_to_desc(raw_use)

        # 표1 셀 업데이트
        if desc_in:
            ws.cell(row=r, column=col_idx_in).value = desc_in
            incoming_counter[desc_in] += 1
        if desc_use:
            ws.cell(row=r, column=col_idx_use).value = desc_use
            usage_counter[desc_use] += 1

    # ----------------- 표4 본문 작성 -----------------
    for i, desc in enumerate(usage_descriptions):
        row = table4_start_row + 1 + i

        # AY: 분류 설명 (첫 행에 단위 캡션 병기)
        c_desc = ws.cell(row=row, column=table4_start_col)
        c_desc.value = f"{desc}" if i == 0 else desc
        c_desc.font = default_font
        c_desc.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c_desc.border = thin_border

        # AZ: 연간입고량 카운트
        c_in = ws.cell(row=row, column=table4_start_col + 1)
        v_in = incoming_counter.get(desc, 0)
        c_in.value = v_in if v_in != 0 else None
        c_in.font = default_font
        c_in.alignment = Alignment(horizontal='center', vertical='center')
        c_in.border = thin_border

        # BA: 연간사용·판매량 카운트
        c_use = ws.cell(row=row, column=table4_start_col + 2)
        v_use = usage_counter.get(desc, 0)
        c_use.value = v_use if v_use != 0 else None
        c_use.font = default_font
        c_use.alignment = Alignment(horizontal='center', vertical='center')
        c_use.border = thin_border
    #=========================================================================#


    # ✅ 저장 및 세션 갱신
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    st.session_state.result_file = output
    st.session_state.processed = True   



# ----------------- 결과 다운로드 -----------------
if st.session_state.processed:
    st.success("✅ 유해성 정보 수집이 완료되었습니다.")
    col1, col2 = st.columns([1, 1])
    
    # 업로드된 파일의 원래 이름
    input_filename = uploaded_file.name

    # 확장자 제거 + '입력파일 ' 접두사 제거
    basename = os.path.splitext(input_filename)[0]
    
    # 출력 파일명 설정
    output_filename = f"{basename}_유해성분석.xlsx"
    
    with col1:
        st.download_button(
            label="📥 결과 엑셀 다운로드",
            data=st.session_state.result_file,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with col2:
        if st.button("🔁 새 파일 업로드"):
            # 🔄 uploader_key를 바꿔야 기존 업로드 상태를 완전히 제거함
            st.session_state.processed = False
            st.session_state.result_file = None
            st.session_state.uploader_key += 1
            st.rerun()
            
            